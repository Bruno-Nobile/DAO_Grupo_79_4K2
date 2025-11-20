#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Módulo de interfaz para reportes
"""

import tkinter as tk
from tkinter import ttk, messagebox, simpledialog, filedialog
import sys
import os

# Agregar directorio padre al path para imports
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from database import get_connection
from config import MATPLOTLIB_AVAILABLE

try:
    from fpdf import FPDF
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

try:
    from openpyxl import Workbook
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
from .ui_utils import enable_treeview_sorting


class ReportesTab(ttk.Frame):
    """Tab para reportes y análisis"""
    
    def __init__(self, container):
        super().__init__(container)
        self.views = {}
        self.periodo_var = tk.StringVar(self, value="mes")
        self.build_ui()

    def build_ui(self):
        """Construye la interfaz de usuario"""
        self.nb = ttk.Notebook(self)
        self.nb.pack(fill=tk.BOTH, expand=True)

        self._build_clientes_section()
        self._build_vehiculos_section()
        self._build_periodos_section()
        self._build_facturacion_section()

    def _build_clientes_section(self):
        frame = ttk.Frame(self.nb)
        self.nb.add(frame, text="Alquiler por cliente")

        controls = ttk.Frame(frame)
        controls.pack(fill=tk.X, padx=5, pady=5)

        ttk.Button(
            controls,
            text="Listar alquileres por cliente",
            command=self.listar_alquileres_por_cliente
        ).pack(side=tk.LEFT)

        btn_pdf = ttk.Button(
            controls,
            text="Exportar listado (PDF)",
            command=lambda: self.exportar_tabla_pdf("clientes", "Alquileres por cliente"),
            state=tk.DISABLED
        )
        btn_pdf.pack(side=tk.LEFT, padx=5)

        tree = self._create_table(frame)
        self.views["clientes"] = {"tree": tree, "columns": [], "rows": [], "buttons": {"pdf": btn_pdf}}

    def _build_vehiculos_section(self):
        frame = ttk.Frame(self.nb)
        self.nb.add(frame, text="Vehículos más alquilados")

        controls = ttk.Frame(frame)
        controls.pack(fill=tk.X, padx=5, pady=5)

        ttk.Button(
            controls,
            text="Mostrar tabla",
            command=self.vehiculos_mas_alquilados
        ).pack(side=tk.LEFT)
        ttk.Button(
            controls,
            text="Gráfico anillo",
            command=self.grafico_vehiculos_anillo
        ).pack(side=tk.LEFT, padx=5)
        ttk.Button(
            controls,
            text="Guardar gráfico",
            command=self.guardar_grafico_vehiculos
        ).pack(side=tk.LEFT, padx=5)

        tree = self._create_table(frame)
        self.views["vehiculos"] = {"tree": tree, "columns": [], "rows": []}

    def _build_periodos_section(self):
        frame = ttk.Frame(self.nb)
        self.nb.add(frame, text="Alquileres por período")

        controls = ttk.Frame(frame)
        controls.pack(fill=tk.X, padx=5, pady=5)

        ttk.Label(controls, text="Período:").pack(side=tk.LEFT)
        ttk.Combobox(
            controls,
            textvariable=self.periodo_var,
            values=["mes", "trimestre", "año"],
            state="readonly",
            width=12
        ).pack(side=tk.LEFT, padx=5)

        ttk.Button(
            controls,
            text="Generar",
            command=self.alquileres_por_periodo
        ).pack(side=tk.LEFT, padx=5)

        btn_pdf = ttk.Button(
            controls,
            text="Exportar PDF",
            command=lambda: self.exportar_tabla_pdf("periodos", "Alquileres por período"),
            state=tk.DISABLED
        )
        btn_pdf.pack(side=tk.LEFT, padx=5)

        btn_excel = ttk.Button(
            controls,
            text="Exportar Excel",
            command=lambda: self.exportar_tabla_excel("periodos"),
            state=tk.DISABLED
        )
        btn_excel.pack(side=tk.LEFT, padx=5)

        ttk.Button(
            controls,
            text="Exportar alquileres (CSV)",
            command=self.exportar_alquileres_csv
        ).pack(side=tk.LEFT, padx=5)

        tree = self._create_table(frame)
        self.views["periodos"] = {
            "tree": tree,
            "columns": [],
            "rows": [],
            "buttons": {"pdf": btn_pdf, "excel": btn_excel}
        }

    def _build_facturacion_section(self):
        frame = ttk.Frame(self.nb)
        self.nb.add(frame, text="Facturación mensual")

        controls = ttk.Frame(frame)
        controls.pack(fill=tk.X, padx=5, pady=5)

        ttk.Button(
            controls,
            text="Ver gráfico",
            command=self.facturacion_mensual
        ).pack(side=tk.LEFT)

        ttk.Button(
            controls,
            text="Guardar gráfico",
            command=self.guardar_grafico_facturacion
        ).pack(side=tk.LEFT, padx=5)

        ttk.Label(
            frame,
            text="Generá el gráfico para visualizar la facturación mensual y guardalo como imagen "
                 "para informes."
        ).pack(fill=tk.X, padx=5, pady=10)

    def _create_table(self, parent):
        container = ttk.Frame(parent)
        container.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        tree = ttk.Treeview(container, show="headings", style="Colored.Treeview")
        vsb = ttk.Scrollbar(container, orient=tk.VERTICAL, command=tree.yview)
        hsb = ttk.Scrollbar(container, orient=tk.HORIZONTAL, command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        container.columnconfigure(0, weight=1)
        container.rowconfigure(0, weight=1)

        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        return tree

    def _show_rows(self, tree, columns, rows):
        """
        Configura la tabla y muestra los datos recibidos.
        columns: lista de tuplas (id_columna, encabezado, ancho)
        rows: lista de iterables con los valores a insertar.
        """
        for child in tree.get_children():
            tree.delete(child)

        col_ids = [c[0] for c in columns]
        tree["columns"] = col_ids

        for cid, heading, width in columns:
            tree.heading(cid, text=heading)
            tree.column(cid, width=width, anchor=tk.W)

        for row in rows:
            tree.insert("", tk.END, values=row)

        if not rows and columns:
            messagebox.showinfo("Información", "No se encontraron datos para mostrar.")

        if columns:
            enable_treeview_sorting(tree)

    def _update_view(self, section, columns, rows):
        view = self.views.get(section)
        if not view:
            return
        self._show_rows(view["tree"], columns, rows)
        view["columns"] = columns
        view["rows"] = rows

        buttons = view.get("buttons", {})
        for key, button in buttons.items():
            if key in ("pdf", "excel"):
                button["state"] = tk.NORMAL if rows else tk.DISABLED

    def listar_alquileres_por_cliente(self):
        """Lista los alquileres por cliente"""
        respuesta = simpledialog.askstring(
            "Listado",
            "ID de cliente (vacío = todos):",
            parent=self
        )

        if respuesta is None:
            return

        respuesta = respuesta.strip()
        cliente_id = None
        if respuesta:
            if not respuesta.isdigit():
                messagebox.showerror("Validación", "Debe ingresar un número entero válido.")
                return
            cliente_id = int(respuesta)

        conn = get_connection()
        c = conn.cursor()
        
        if cliente_id:
            c.execute("""SELECT a.*, c.apellido||', '||c.nombre as cliente,
                                v.patente || ' - ' || v.marca || ' ' || v.modelo as vehiculo
                         FROM alquiler a
                         JOIN cliente c ON a.id_cliente=c.id_cliente
                         JOIN vehiculo v ON a.id_vehiculo=v.id_vehiculo
                         WHERE a.id_cliente=?
                         ORDER BY a.fecha_inicio DESC""", (cliente_id,))
        else:
            c.execute("""SELECT a.*, c.apellido||', '||c.nombre as cliente,
                                v.patente || ' - ' || v.marca || ' ' || v.modelo as vehiculo
                         FROM alquiler a
                         JOIN cliente c ON a.id_cliente=c.id_cliente
                         JOIN vehiculo v ON a.id_vehiculo=v.id_vehiculo
                         ORDER BY a.fecha_inicio DESC""")
        
        rows = c.fetchall()
        conn.close()
        
        datos = [
            (
                r["id_alquiler"],
                r["cliente"],
                r["vehiculo"],
                r["fecha_inicio"],
                r["fecha_fin"],
                r["costo_total"]
            )
            for r in rows
        ]

        columnas = [
            ("id_alquiler", "ID", 60),
            ("cliente", "Cliente", 200),
            ("vehiculo", "Vehículo", 220),
            ("fecha_inicio", "Inicio", 110),
            ("fecha_fin", "Fin", 110),
            ("costo_total", "Costo ($)", 100)
        ]

        self._update_view("clientes", columnas, datos)

    def vehiculos_mas_alquilados(self):
        """Lista los vehículos más alquilados"""
        conn = get_connection()
        c = conn.cursor()
        
        c.execute("""SELECT v.id_vehiculo, v.patente, v.marca||' '||v.modelo as desc, COUNT(a.id_alquiler) as veces
                     FROM vehiculo v
                     LEFT JOIN alquiler a ON v.id_vehiculo = a.id_vehiculo
                     GROUP BY v.id_vehiculo
                     ORDER BY veces DESC""")
        
        rows = c.fetchall()
        conn.close()

        datos = [
            (
                r["id_vehiculo"],
                r["patente"],
                r["desc"],
                r["veces"]
            )
            for r in rows
        ]

        columnas = [
            ("id_vehiculo", "ID", 60),
            ("patente", "Patente", 120),
            ("desc", "Descripción", 240),
            ("veces", "Cant. Alquileres", 140)
        ]

        self._update_view("vehiculos", columnas, datos)

    def grafico_vehiculos_anillo(self, save_path=None):
        """Genera gráfico de anillo para vehículos más alquilados"""
        if not MATPLOTLIB_AVAILABLE:
            messagebox.showerror("Error", "matplotlib es requerido para los gráficos.")
            return

        import matplotlib.pyplot as plt

        conn = get_connection()
        c = conn.cursor()
        c.execute("""SELECT v.patente || ' - ' || v.marca||' '||v.modelo as desc, COUNT(a.id_alquiler) as veces
                     FROM vehiculo v
                     LEFT JOIN alquiler a ON v.id_vehiculo = a.id_vehiculo
                     GROUP BY v.id_vehiculo
                     ORDER BY veces DESC""")
        rows = c.fetchall()
        conn.close()

        # Filtrar vehículos con 0 alquileres
        datos_filtrados = [(r["desc"], r["veces"]) for r in rows if r["veces"] > 0]
        
        if not datos_filtrados:
            messagebox.showinfo("Información", "No hay vehículos con alquileres para mostrar en el gráfico.")
            return
        
        labels = [d[0] for d in datos_filtrados]
        valores = [d[1] for d in datos_filtrados]

        # Paleta de colores amplia para evitar repeticiones
        # Usamos varios colormaps combinados para tener más variedad
        import matplotlib.cm as cm
        
        n_colores = len(valores)
        colores = []
        
        # Combinar múltiples colormaps para tener una paleta amplia
        # tab20 tiene 20 colores distintos, Set3 tiene 12, Pastel1 y Pastel2 tienen 8 cada uno
        cmap1 = plt.get_cmap('tab20')
        cmap2 = plt.get_cmap('Set3')
        cmap3 = plt.get_cmap('Pastel1')
        cmap4 = plt.get_cmap('Pastel2')
        cmap5 = plt.get_cmap('Set1')
        cmap6 = plt.get_cmap('Set2')
        
        # Generar colores de diferentes colormaps para máxima variedad
        for i in range(n_colores):
            if i < 20:
                # Usar índices discretos para tab20 (tiene exactamente 20 colores)
                colores.append(cmap1(i % 20))
            elif i < 32:
                colores.append(cmap2((i - 20) % 12))
            elif i < 40:
                colores.append(cmap3((i - 32) % 8))
            elif i < 48:
                colores.append(cmap4((i - 40) % 8))
            elif i < 57:
                colores.append(cmap5((i - 48) % 9))
            else:
                colores.append(cmap6((i - 57) % 8))

        fig, ax = plt.subplots(figsize=(6, 5))
        wedges, texts, autotexts = ax.pie(valores, labels=labels, startangle=90, 
                                         wedgeprops=dict(width=0.4), autopct='%1.1f%%',
                                         colors=colores)
        ax.set_aspect('equal')
        ax.set_title("Vehículos más alquilados (anillo)")

        if save_path:
            fig.savefig(save_path, dpi=150, bbox_inches='tight')
            plt.close(fig)
            messagebox.showinfo("Exportar", f"Gráfico guardado en {save_path}")
        else:
            plt.show()

    def guardar_grafico_vehiculos(self):
        """Permite guardar el gráfico de anillo como imagen"""
        if not MATPLOTLIB_AVAILABLE:
            messagebox.showerror("Error", "matplotlib es requerido para los gráficos.")
            return
        filename = filedialog.asksaveasfilename(
            defaultextension=".png",
            filetypes=[("PNG", "*.png"), ("Imagen JPG", "*.jpg"), ("Todos", "*.*")]
        )
        if not filename:
            return
        self.grafico_vehiculos_anillo(save_path=filename)

    def facturacion_mensual(self):
        """Muestra un gráfico de facturación mensual"""
        if not MATPLOTLIB_AVAILABLE:
            messagebox.showerror(
                "Error",
                "matplotlib no está instalado. Instale con 'pip install matplotlib' para ver gráficos."
            )
            return
        
        import matplotlib.pyplot as plt
        
        conn = get_connection()
        c = conn.cursor()
        
        c.execute("""SELECT substr(fecha_inicio,1,7) as mes, SUM(costo_total) as total
                     FROM alquiler
                     GROUP BY mes
                     ORDER BY mes""")
        
        rows = c.fetchall()
        conn.close()
        
        meses = [r["mes"] for r in rows]
        tot = [r["total"] for r in rows]
        
        if not meses:
            messagebox.showinfo("Info", "No hay datos de facturación")
            return
        
        plt.figure(figsize=(8, 4))
        plt.bar(meses, tot)
        plt.title("Facturación mensual")
        plt.xlabel("Mes")
        plt.ylabel("Total ($)")
        plt.tight_layout()
        plt.show()

    def guardar_grafico_facturacion(self):
        """Guarda el gráfico de facturación en imagen"""
        if not MATPLOTLIB_AVAILABLE:
            messagebox.showerror("Error", "matplotlib es requerido para exportar gráficos.")
            return

        filename = filedialog.asksaveasfilename(
            defaultextension=".png",
            filetypes=[("PNG", "*.png"), ("Imagen JPG", "*.jpg"), ("Todos", "*.*")]
        )
        if not filename:
            return

        import matplotlib.pyplot as plt

        conn = get_connection()
        c = conn.cursor()
        c.execute("""SELECT substr(fecha_inicio,1,7) as mes, SUM(costo_total) as total
                     FROM alquiler
                     GROUP BY mes
                     ORDER BY mes""")
        rows = c.fetchall()
        conn.close()

        meses = [r["mes"] for r in rows]
        tot = [r["total"] for r in rows]
        if not meses:
            messagebox.showinfo("Info", "No hay datos de facturación")
            return

        fig, ax = plt.subplots(figsize=(8, 4))
        ax.bar(meses, tot)
        ax.set_title("Facturación mensual")
        ax.set_xlabel("Mes")
        ax.set_ylabel("Total ($)")
        fig.tight_layout()
        fig.savefig(filename, dpi=150, bbox_inches='tight')
        plt.close(fig)
        messagebox.showinfo("Exportar", f"Gráfico guardado en {filename}")

    def exportar_alquileres_csv(self):
        """Exporta la lista de alquileres a un archivo CSV"""
        import csv
        
        fname = "alquileres_export.csv"
        conn = get_connection()
        c = conn.cursor()
        
        c.execute("""SELECT a.id_alquiler, a.fecha_inicio, a.fecha_fin, a.costo_total,
                            c.nombre||' '||c.apellido as cliente, v.patente as vehiculo
                     FROM alquiler a
                     JOIN cliente c ON a.id_cliente=c.id_cliente
                     JOIN vehiculo v ON a.id_vehiculo=v.id_vehiculo
                     ORDER BY a.fecha_inicio DESC""")
        
        rows = c.fetchall()
        conn.close()
        
        with open(fname, "w", newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(["id_alquiler", "fecha_inicio", "fecha_fin", "costo_total", "cliente", "vehiculo"])
            for r in rows:
                writer.writerow([
                    r["id_alquiler"], r["fecha_inicio"], r["fecha_fin"],
                    r["costo_total"], r["cliente"], r["vehiculo"]
                ])
        
        messagebox.showinfo("Exportar", f"Exportado a {fname}")

    def alquileres_por_periodo(self):
        """Genera un resumen de alquileres agrupado por período."""
        periodo = self.periodo_var.get()
        if periodo not in ("mes", "trimestre", "año"):
            messagebox.showerror("Error", "Seleccione un período válido.")
            return

        conn = get_connection()
        c = conn.cursor()

        if periodo == "mes":
            campo = "substr(fecha_inicio,1,7)"
            titulo = "Alquileres por mes"
        elif periodo == "trimestre":
            campo = ("substr(fecha_inicio,1,4) || '-T' || ("
                     "CAST(((CAST(substr(fecha_inicio,6,2) AS INTEGER)-1) / 3) AS INTEGER) + 1)")
            titulo = "Alquileres por trimestre"
        else:
            campo = "substr(fecha_inicio,1,4)"
            titulo = "Alquileres por año"

        c.execute(f"""SELECT {campo} as periodo,
                             COUNT(*) as cantidad,
                             SUM(costo_total) as total
                      FROM alquiler
                      GROUP BY periodo
                      ORDER BY periodo""")

        rows = c.fetchall()
        conn.close()

        datos = [(r["periodo"], r["cantidad"], r["total"]) for r in rows]
        columnas = [
            ("periodo", "Período", 140),
            ("cantidad", "Cantidad", 120),
            ("total", "Total facturado ($)", 180)
        ]
        self._update_view("periodos", columnas, datos)
        if not datos:
            messagebox.showinfo("Información", "No hay datos para el período seleccionado.")

    def exportar_tabla_pdf(self, section, titulo):
        """Genera un PDF con la tabla actual."""
        if not PDF_AVAILABLE:
            messagebox.showerror("Error", "Instale la librería fpdf2 para exportar a PDF.")
            return
        
        view = self.views.get(section)
        if not view or not view["rows"]:
            messagebox.showwarning("Información", "No hay datos para exportar.")
            return

        filename = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF", "*.pdf")],
            title="Guardar reporte PDF"
        )
        if not filename:
            return

        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.add_page()
        pdf.set_font("Helvetica", "B", 14)
        pdf.cell(0, 10, titulo, ln=True, align="C")
        pdf.ln(5)

        pdf.set_font("Helvetica", "B", 10)
        col_widths = []
        for _, heading, _ in view["columns"]:
            width = max(30, pdf.get_string_width(heading) + 10)
            col_widths.append(width)
            pdf.cell(width, 8, heading, border=1, align="C")
        pdf.ln()

        pdf.set_font("Helvetica", "", 9)
        for row in view["rows"]:
            for idx, value in enumerate(row):
                pdf.cell(col_widths[idx], 8, str(value), border=1)
            pdf.ln()

        pdf.output(filename)
        messagebox.showinfo("Exportar", f"Reporte guardado en {filename}")

    def exportar_tabla_excel(self, section):
        """Exporta la tabla actual a Excel."""
        if not EXCEL_AVAILABLE:
            messagebox.showerror("Error", "Instale la librería openpyxl para exportar a Excel.")
            return
        
        view = self.views.get(section)
        if not view or not view["rows"]:
            messagebox.showwarning("Información", "No hay datos para exportar.")
            return

        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            title="Guardar reporte Excel"
        )
        if not filename:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte"

        headers = [col[1] for col in view["columns"]]
        ws.append(headers)
        for row in view["rows"]:
            ws.append(list(row))

        wb.save(filename)
        messagebox.showinfo("Exportar", f"Reporte guardado en {filename}")
