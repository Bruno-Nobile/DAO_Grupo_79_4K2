#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Servicio de Reportes - Lógica de negocio para reportes
Programación Orientada a Objetos - Encapsula la lógica de reportes
Programación Funcional - Uso de funciones para procesar datos
"""

from persistence.database_connection import DatabaseConnection
from datetime import datetime, date
import base64
import io
from config import MATPLOTLIB_AVAILABLE

if MATPLOTLIB_AVAILABLE:
    import matplotlib
    matplotlib.use('Agg')  # Backend no interactivo para servidor
    import matplotlib.pyplot as plt
    from matplotlib.ticker import FuncFormatter

# Verificar disponibilidad de openpyxl para exportación a Excel
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ReportesService:
    """
    Programación Orientada a Objetos - Servicio de reportes
    Encapsula la lógica de generación de reportes
    """
    
    def __init__(self):
        """
        Programación Orientada a Objetos - Constructor
        Patrón Singleton - Uso de DatabaseConnection
        """
        # Patrón Singleton - Obtener instancia única de conexión
        self._db = DatabaseConnection()
    
    def alquileres_por_cliente(self):
        """
        Reporte: Listado de alquileres por cliente
        Programación Estructurada - Función bien organizada
        """
        query = """
        SELECT 
            c.id_cliente,
            c.nombre || ' ' || c.apellido as cliente_nombre,
            COUNT(a.id_alquiler) as total_alquileres,
            SUM(a.costo_total) as total_facturado
        FROM cliente c
        LEFT JOIN alquiler a ON c.id_cliente = a.id_cliente
        GROUP BY c.id_cliente, c.nombre, c.apellido
        ORDER BY total_alquileres DESC, cliente_nombre
        """
        cursor = self._db.execute_query(query)
        rows = cursor.fetchall()
        
        # Programación Funcional - Transformar filas a diccionarios
        return [dict(row) for row in rows]
    
    def detalle_alquileres_por_cliente(self, id_cliente):
        """
        Reporte: Detalle de alquileres de un cliente específico
        Programación Estructurada - Función bien organizada
        """
        query = """
        SELECT 
            a.id_alquiler,
            a.fecha_inicio,
            a.fecha_fin,
            a.costo_total,
            v.patente || ' - ' || v.marca || ' ' || v.modelo as vehiculo,
            e.nombre || ' ' || e.apellido as empleado
        FROM alquiler a
        JOIN vehiculo v ON a.id_vehiculo = v.id_vehiculo
        LEFT JOIN empleado e ON a.id_empleado = e.id_empleado
        WHERE a.id_cliente = ?
        ORDER BY a.fecha_inicio DESC
        """
        cursor = self._db.execute_query(query, (id_cliente,))
        rows = cursor.fetchall()
        
        # Programación Funcional - Transformar filas a diccionarios
        return [dict(row) for row in rows]
    
    def vehiculos_mas_alquilados(self):
        """
        Reporte: Vehículos más alquilados
        Programación Estructurada - Función bien organizada
        """
        query = """
        SELECT 
            v.id_vehiculo,
            v.patente,
            v.marca || ' ' || v.modelo as descripcion,
            COUNT(a.id_alquiler) as veces_alquilado,
            SUM(a.costo_total) as total_facturado
        FROM vehiculo v
        LEFT JOIN alquiler a ON v.id_vehiculo = a.id_vehiculo
        GROUP BY v.id_vehiculo, v.patente, v.marca, v.modelo
        ORDER BY veces_alquilado DESC, total_facturado DESC
        """
        cursor = self._db.execute_query(query)
        rows = cursor.fetchall()
        
        # Programación Funcional - Transformar filas a diccionarios
        return [dict(row) for row in rows]
    
    def alquileres_por_periodo(self, periodo='mes'):
        """
        Reporte: Alquileres por período (mes, trimestre, año)
        Programación Estructurada - Función bien organizada
        Programación Funcional - Lógica condicional para diferentes períodos
        """
        if periodo == 'mes':
            # Agrupar por mes
            query = """
            SELECT 
                strftime('%Y-%m', fecha_inicio) as periodo,
                COUNT(*) as cantidad_alquileres,
                SUM(costo_total) as total_facturado
            FROM alquiler
            GROUP BY strftime('%Y-%m', fecha_inicio)
            ORDER BY periodo DESC
            """
        elif periodo == 'trimestre':
            # Agrupar por trimestre
            query = """
            SELECT 
                strftime('%Y', fecha_inicio) || '-Q' || 
                CAST((CAST(strftime('%m', fecha_inicio) AS INTEGER) - 1) / 3 + 1 AS TEXT) as periodo,
                COUNT(*) as cantidad_alquileres,
                SUM(costo_total) as total_facturado
            FROM alquiler
            GROUP BY strftime('%Y', fecha_inicio), 
                     CAST((CAST(strftime('%m', fecha_inicio) AS INTEGER) - 1) / 3 + 1 AS INTEGER)
            ORDER BY periodo DESC
            """
        elif periodo == 'año':
            # Agrupar por año
            query = """
            SELECT 
                strftime('%Y', fecha_inicio) as periodo,
                COUNT(*) as cantidad_alquileres,
                SUM(costo_total) as total_facturado
            FROM alquiler
            GROUP BY strftime('%Y', fecha_inicio)
            ORDER BY periodo DESC
            """
        else:
            return []
        
        cursor = self._db.execute_query(query)
        rows = cursor.fetchall()
        
        # Programación Funcional - Transformar filas a diccionarios
        return [dict(row) for row in rows]
    
    def facturacion_mensual_grafico(self):
        """
        Reporte: Facturación mensual en gráfico de barras
        Programación Orientada a Objetos - Método de comportamiento
        Programación Funcional - Procesamiento de datos para gráfico
        """
        if not MATPLOTLIB_AVAILABLE:
            return None
        
        # Obtener datos mensuales
        datos = self.alquileres_por_periodo('mes')
        
        if not datos:
            return None
        
        # Programación Funcional - Extraer períodos y totales usando map
        periodos = list(map(lambda d: d['periodo'], datos))
        totales = list(map(lambda d: float(d['total_facturado'] or 0), datos))
        
        # Función para formatear números en el eje Y
        def formatear_valor(val, pos):
            """Formatea valores grandes en formato legible (K, M)"""
            if val >= 1_000_000:
                return f'${val/1_000_000:.2f}M'
            elif val >= 1_000:
                return f'${val/1_000:.1f}K'
            else:
                return f'${val:.0f}'
        
        # Crear gráfico
        fig, ax = plt.subplots(figsize=(12, 6))
        bars = ax.bar(periodos, totales, color='steelblue', alpha=0.7, edgecolor='navy', linewidth=1.2)
        
        # Configurar ejes
        ax.set_xlabel('Mes (YYYY-MM)', fontsize=12, fontweight='bold')
        ax.set_ylabel('Facturación', fontsize=12, fontweight='bold')
        ax.set_title('Facturación Mensual', fontsize=16, fontweight='bold', pad=20)
        
        # Formatear eje Y con formato legible
        ax.yaxis.set_major_formatter(FuncFormatter(formatear_valor))
        
        # Rotar etiquetas del eje X
        plt.xticks(rotation=45, ha='right')
        
        # Agregar grid
        ax.grid(axis='y', alpha=0.3, linestyle='--')
        ax.set_axisbelow(True)
        
        # Agregar valores en las barras si hay espacio suficiente
        max_val = max(totales) if totales else 0
        if max_val > 0:
            for i, (bar, total) in enumerate(zip(bars, totales)):
                height = bar.get_height()
                # Mostrar valor solo si la barra es lo suficientemente alta
                if height > max_val * 0.05:  # Solo si la barra es > 5% del máximo
                    if total >= 1_000_000:
                        label = f'${total/1_000_000:.2f}M'
                    elif total >= 1_000:
                        label = f'${total/1_000:.1f}K'
                    else:
                        label = f'${total:.0f}'
                    ax.text(bar.get_x() + bar.get_width()/2., height,
                           label,
                           ha='center', va='bottom', fontsize=9, fontweight='bold')
        
        # Ajustar layout
        plt.tight_layout()
        
        # Convertir gráfico a imagen base64
        img_buffer = io.BytesIO()
        plt.savefig(img_buffer, format='png', dpi=100, bbox_inches='tight')
        img_buffer.seek(0)
        img_base64 = base64.b64encode(img_buffer.getvalue()).decode('utf-8')
        plt.close()
        
        return img_base64
    
    def exportar_vehiculos_mas_alquilados_excel(self):
        """
        Exporta vehículos más alquilados a archivo Excel
        Programación Orientada a Objetos - Método de comportamiento
        Programación Estructurada - Función bien organizada
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl no está instalado. Instálelo con: pip install openpyxl")
        
        # Obtener datos
        datos = self.vehiculos_mas_alquilados()
        
        # Crear libro de trabajo Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Vehículos Más Alquilados"
        
        # Estilos para el encabezado
        # Programación Orientada a Objetos - Uso de objetos de estilo
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Encabezados
        headers = ["ID Vehículo", "Patente", "Descripción", "Veces Alquilado", "Total Facturado ($)"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 20
        
        # Programación Funcional - Usar map para procesar datos
        # Agregar datos
        for row_num, vehiculo in enumerate(datos, 2):
            ws.cell(row=row_num, column=1, value=vehiculo.get('id_vehiculo', ''))
            ws.cell(row=row_num, column=2, value=vehiculo.get('patente', ''))
            ws.cell(row=row_num, column=3, value=vehiculo.get('descripcion', ''))
            ws.cell(row=row_num, column=4, value=vehiculo.get('veces_alquilado', 0))
            
            # Formatear total facturado como número con 2 decimales
            total = float(vehiculo.get('total_facturado', 0) or 0)
            cell_total = ws.cell(row=row_num, column=5, value=total)
            cell_total.number_format = '#,##0.00'
            
            # Aplicar bordes a todas las celdas
            for col_num in range(1, 6):
                ws.cell(row=row_num, column=col_num).border = border
        
        # Agregar fila de totales
        if datos:
            total_row = len(datos) + 3
            ws.cell(row=total_row, column=3, value="TOTALES:").font = Font(bold=True)
            ws.cell(row=total_row, column=4, 
                   value=f"=SUM(D2:D{len(datos)+1})").font = Font(bold=True)
            ws.cell(row=total_row, column=5, 
                   value=f"=SUM(E2:E{len(datos)+1})").font = Font(bold=True)
            ws.cell(row=total_row, column=5).number_format = '#,##0.00'
        
        # Agregar información adicional
        info_row = len(datos) + 5
        ws.cell(row=info_row, column=1, value="Fecha de generación:").font = Font(italic=True)
        ws.cell(row=info_row, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        
        # Guardar en buffer de memoria
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer
    
    def exportar_alquileres_por_cliente_excel(self):
        """
        Exporta alquileres por cliente a archivo Excel
        Programación Orientada a Objetos - Método de comportamiento
        Programación Estructurada - Función bien organizada
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl no está instalado. Instálelo con: pip install openpyxl")
        
        # Obtener datos
        datos = self.alquileres_por_cliente()
        
        # Crear libro de trabajo Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Alquileres por Cliente"
        
        # Estilos para el encabezado
        # Programación Orientada a Objetos - Uso de objetos de estilo
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Encabezados
        headers = ["ID Cliente", "Cliente", "Total Alquileres", "Total Facturado ($)"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 20
        
        # Programación Funcional - Usar enumerate para procesar datos
        # Agregar datos
        for row_num, cliente in enumerate(datos, 2):
            ws.cell(row=row_num, column=1, value=cliente.get('id_cliente', ''))
            ws.cell(row=row_num, column=2, value=cliente.get('cliente_nombre', ''))
            ws.cell(row=row_num, column=3, value=cliente.get('total_alquileres', 0))
            
            # Formatear total facturado como número con 2 decimales
            total = float(cliente.get('total_facturado', 0) or 0)
            cell_total = ws.cell(row=row_num, column=4, value=total)
            cell_total.number_format = '#,##0.00'
            
            # Aplicar bordes a todas las celdas
            for col_num in range(1, 5):
                ws.cell(row=row_num, column=col_num).border = border
        
        # Agregar fila de totales
        if datos:
            total_row = len(datos) + 3
            ws.cell(row=total_row, column=2, value="TOTALES:").font = Font(bold=True)
            ws.cell(row=total_row, column=3, 
                   value=f"=SUM(C2:C{len(datos)+1})").font = Font(bold=True)
            ws.cell(row=total_row, column=4, 
                   value=f"=SUM(D2:D{len(datos)+1})").font = Font(bold=True)
            ws.cell(row=total_row, column=4).number_format = '#,##0.00'
        
        # Agregar información adicional
        info_row = len(datos) + 5
        ws.cell(row=info_row, column=1, value="Fecha de generación:").font = Font(italic=True)
        ws.cell(row=info_row, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        
        # Guardar en buffer de memoria
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer

